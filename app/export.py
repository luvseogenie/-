"""엑셀 내려받기."""
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import config

COLUMNS = [
    ("판정", "verdict_label", 12), ("상품명", "name", 50), ("상품ID", "product_id", 12),
    ("카테고리", "category_path", 40), ("판매가격", "price", 11), ("실제확인가", "verified_price", 11),
    ("리뷰", "review_count", 8), ("평점", "rating", 6), ("28일 판매", "sales_28", 10), ("일평균", "daily_avg", 8),
    ("전환율(%)", "conversion", 9), ("28일 조회", "views_28", 10), ("28일 매출", "revenue_28", 14),
    ("리뷰당 판매", "sales_per_review", 10), ("배송", "delivery", 14), ("광고", "is_ad", 6), ("품절", "sold_out", 6),
    ("못 파는 물건", "restricted", 12), ("비슷한 판매자", "seller_count", 10), ("옵션 수", "option_count", 8),
    ("링크", "url", 60),
]


def build_xlsx(rows: list[dict], title="소싱 결과") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    head_fill = PatternFill("solid", fgColor="1F2432")
    head_font = Font(bold=True, color="FFFFFF")
    for i, (label, _, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=label)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    for r, row in enumerate(rows, 2):
        for i, (_, key, _) in enumerate(COLUMNS, 1):
            v = row.get(key)
            if isinstance(v, bool):
                v = "예" if v else ""
            if key in ("is_ad", "sold_out") and v in (0, 1):
                v = "예" if v else ""
            ws.cell(row=r, column=i, value=v)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 1)}"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = config.EXPORT_DIR / f"쿠팡소싱_{stamp}.xlsx"
    wb.save(path)
    return str(path)
