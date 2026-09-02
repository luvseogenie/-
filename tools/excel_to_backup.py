"""예전 '광고계산기' 엑셀을 확장 프로그램 백업 JSON 으로 바꾼다.

기본: 4번 시트(광고 장부확인)의 확정 값을 캠페인×날짜 legacy 로, 1번 시트의 매핑을 옵션·마진(엑셀 마지막 날 다음 날부터 적용)으로.
--raw 를 주면 예전 방식(2·3번 시트의 원자료)으로 바꾼다.

  python tools/excel_to_backup.py 광고계산기.xlsx 백업.json
→ 확장 프로그램: 장부 보기 → 백업 · 설정 → 복원 에서 JSON 선택
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("원", "")
    pct = s.endswith("%")
    try:
        n = float(s.rstrip("%"))
    except ValueError:
        return 0.0
    return n / 100 if pct else n


def _pct(v):
    """분수로 통일 (0.078). 문자열 '7.8%' → 0.078, 숫자 7.8 → 0.078, 0.078 → 0.078"""
    if isinstance(v, str) and v.strip().endswith("%"):
        return _num(v)
    n = _num(v)
    return n / 100 if n > 1 else n


def _iso(v):
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _id(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def _find_sheet(wb, *keys):
    for ws in wb.worksheets:
        if all(k in ws.title for k in keys):
            return ws
    return None


def _header_row(ws, *keys, limit=10):
    for r in range(1, limit + 1):
        vals = [str(ws.cell(r, c).value or "").replace("\n", "") for c in range(1, ws.max_column + 1)]
        # 머리글은 짧은 칸들이다. 긴 설명문(2행 메모)이 걸리지 않도록 20자 이하 칸만 본다.
        if all(any(k in v and len(v) <= 20 for v in vals) for k in keys) and sum(1 for v in vals if v) >= 3:
            return r, vals
    return None, None


def _col(vals, *names):
    for i, v in enumerate(vals):
        if any(v.replace(" ", "").startswith(n.replace(" ", "")) for n in names):
            return i + 1
    return None


def convert_raw(xlsx: Path) -> dict:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    data = {"options": [], "margins": [], "sales": {}, "ads": {}}

    ws = _find_sheet(wb, "매핑") or wb.worksheets[1]
    hr, vals = _header_row(ws, "옵션", "캠페인")
    c_name, c_id, c_camp, c_margin = _col(vals, "상품명"), _col(vals, "옵션 ID", "옵션ID"), _col(vals, "광고 캠페인", "캠페인"), _col(vals, "옵션마진", "마진")
    seen = set()
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        oid = _id(row[c_id - 1]) if c_id else ""
        if not oid or not oid.isdigit() or oid in seen:
            continue
        seen.add(oid)
        data["options"].append({"option_id": oid, "product_name": str(row[c_name - 1] or "").strip(), "campaign": str(row[c_camp - 1] or "").strip(), "sort_order": len(seen)})
        m = _num(row[c_margin - 1]) if c_margin else 0
        if m:
            data["margins"].append({"option_id": oid, "effective_from": "", "margin": m, "note": "엑셀에서 가져옴"})

    ws = _find_sheet(wb, "광고 실적") or wb.worksheets[2]
    hr, vals = _header_row(ws, "캠페인", "날짜")
    c = {k: _col(vals, *n) for k, n in {
        "campaign": ["캠페인"], "date": ["날짜"], "target": ["목표효율"], "budget": ["광고예산"], "spend": ["집행광고비", "집행 광고비"],
        "rev": ["광고전환매출", "광고전환 매출"], "conv": ["전환율"], "ctr": ["클릭률"], "imp": ["노출수"], "clk": ["클릭수"], "orders": ["광고전환 판매수", "광고전환판매수"], "action": ["ACTION"],
    }.items()}
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        camp = str(row[c["campaign"] - 1] or "").strip() if c["campaign"] else ""
        date = _iso(row[c["date"] - 1]) if c["date"] else None
        if not camp or not date:
            continue
        g = lambda k: row[c[k] - 1] if c[k] else None  # noqa: E731
        data["ads"].setdefault(date, {})[camp] = {
            "date": date, "campaign": camp, "target_roas": _num(g("target")), "budget": _num(g("budget")), "spend": _num(g("spend")),
            "ad_revenue": _num(g("rev")), "conversion": _pct(g("conv")), "ctr": _pct(g("ctr")), "impressions": _num(g("imp")),
            "clicks": _num(g("clk")), "ad_orders": _num(g("orders")), "action": str(g("action") or "").strip(),
        }

    ws = _find_sheet(wb, "매출 실적") or wb.worksheets[3]
    hr, vals = _header_row(ws, "옵션", "매출")
    c = {k: _col(vals, *n) for k, n in {
        "date": ["날짜"], "oid": ["옵션ID", "옵션 ID"], "oname": ["옵션명"], "pname": ["상품명"], "pid": ["등록상품ID"], "cat": ["카테고리"], "type": ["판매방식"],
        "rev": ["매출"], "orders": ["주문"], "qty": ["판매량"], "vis": ["방문자"], "views": ["조회"], "carts": ["장바구니"], "conv": ["구매전환율"],
    }.items()}
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        oid = _id(row[c["oid"] - 1]) if c["oid"] else ""
        date = _iso(row[c["date"] - 1]) if c["date"] else None
        if not oid or not oid.isdigit() or not date:
            continue
        g = lambda k: row[c[k] - 1] if c[k] else None  # noqa: E731
        data["sales"].setdefault(date, {})[oid] = {
            "date": date, "option_id": oid, "option_name": str(g("oname") or "").strip(), "product_name": str(g("pname") or "").strip(),
            "product_id": _id(g("pid")), "category": str(g("cat") or "").strip(), "sales_type": str(g("type") or "").strip(),
            "revenue": _num(g("rev")), "orders": _num(g("orders")), "quantity": _num(g("qty")), "visitors": _num(g("vis")),
            "views": _num(g("views")), "carts": _num(g("carts")), "conversion": _pct(g("conv")) if g("conv") not in (None, "") else None,
        }
        if oid not in seen:  # 매핑에 없는 옵션은 이름만 등록
            seen.add(oid)
            data["options"].append({"option_id": oid, "product_name": str(g("oname") or g("pname") or "").strip(), "campaign": "", "sort_order": len(seen)})
    wb.close()
    return data


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    src = Path(argv[1]); dst = Path(argv[2]) if len(argv) > 2 else src.with_suffix(".backup.json")
    data = convert(src)
    dst.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    n_sales = sum(len(v) for v in data["sales"].values()); n_ads = sum(len(v) for v in data["ads"].values())
    dates = sorted(set(data["sales"]) | set(data["ads"]))
    print(f"옵션 {len(data['options'])}개, 마진 {len(data['margins'])}개, 판매 {n_sales}행, 광고 {n_ads}행, 기간 {dates[0]}~{dates[-1]} → {dst} ({dst.stat().st_size/1e6:.1f} MB)")
    return 0


LABELS = [("target_roas", ["목표효율"]), ("roas", ["광고수익률"]), ("budget", ["광고예산"]), ("spend_vat", ["집행광고비*10%", "집행광고비"]),
          ("cpc", ["cpc"]), ("impressions", ["노출수"]), ("ctr", ["클릭률"]), ("conversion", ["전환율"]), ("ad_orders", ["광고전환판매수"]),
          ("organic_qty", ["자연판매수"]), ("actual_qty", ["실제판매수"]), ("margin_total", ["광고마진", "판매마진"]), ("profit", ["순이익"])]


def _label_field(label):
    n = str(label or "").replace(" ", "").replace("\n", "").lower()
    for f, keys in LABELS:
        if any(n.startswith(k.lower()) for k in keys):
            return f
    return None


def convert(xlsx: Path) -> dict:
    """4번 시트 기준 변환 (기본)."""
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = _find_sheet(wb, "장부") or wb.worksheets[4]
    rows = list(ws.iter_rows(values_only=True))
    hr, date_cols = None, []
    for r in range(min(10, len(rows))):
        cols = [(i, _iso(v)) for i, v in enumerate(rows[r]) if _iso(v)]
        if len(cols) >= 5:
            hr, date_cols = r, cols
            break
    if hr is None:
        raise SystemExit("4번 시트에서 날짜 행을 찾지 못했습니다")
    legacy, camp = {}, None
    for row in rows[hr + 1:]:
        a = str(row[0] or "").strip(); b = str(row[1] or "").strip()
        if a and "이곳에" not in a and "입력해" not in a:
            camp = a
        elif a:
            camp = None
        if not camp or not b:
            continue
        f = _label_field(b)
        if not f:
            continue
        for ci, date in date_cols:
            v = row[ci] if ci < len(row) else None
            if v in (None, ""):
                continue
            n = float(v) if isinstance(v, (int, float)) else _num(v)
            legacy.setdefault(date, {}).setdefault(camp, {})[f] = n
    for date in list(legacy):
        for c in list(legacy[date]):
            L = legacy[date][c]
            active = L.get("spend_vat", 0) > 0 or L.get("impressions", 0) > 0 or L.get("ad_orders", 0) > 0 or L.get("actual_qty", 0) > 0 or L.get("margin_total", 0) != 0
            if not active:
                del legacy[date][c]; continue
            L["spend"] = L.get("spend_vat", 0) / 1.1; L["ad_revenue"] = L.get("roas", 0) * L["spend"]; L["clicks"] = L["spend"] / L["cpc"] if L.get("cpc") else 0
            if "actual_qty" not in L and "organic_qty" in L:
                L["actual_qty"] = L.get("ad_orders", 0) + L["organic_qty"]
            L.setdefault("profit", L.get("margin_total", 0) - L.get("spend_vat", 0))
            L["src"] = "excel"
        if not legacy[date]:
            del legacy[date]
    dates = sorted(legacy)
    last = dt.date.fromisoformat(dates[-1]); margin_from = (last + dt.timedelta(days=1)).isoformat()
    data = {"options": [], "margins": [], "sales": {}, "ads": {}, "legacy": legacy, "imports": []}
    ws = _find_sheet(wb, "매핑") or wb.worksheets[1]
    hr2, vals = _header_row(ws, "옵션", "캠페인")
    if hr2:
        c_name, c_id, c_camp, c_margin = _col(vals, "상품명"), _col(vals, "옵션 ID", "옵션ID"), _col(vals, "광고 캠페인", "캠페인"), _col(vals, "옵션마진", "마진")
        seen = set()
        for row in ws.iter_rows(min_row=hr2 + 1, values_only=True):
            oid = _id(row[c_id - 1]) if c_id else ""
            if not oid.isdigit() or oid in seen:
                continue
            seen.add(oid)
            data["options"].append({"option_id": oid, "product_name": str(row[c_name - 1] or "").strip(), "campaign": str(row[c_camp - 1] or "").strip(), "sort_order": len(seen)})
            m = _num(row[c_margin - 1]) if c_margin else 0
            if m:
                data["margins"].append({"option_id": oid, "effective_from": "", "margin": m, "note": "엑셀에서 가져옴"})
    wb.close()
    return data


def _main_wrapper(argv):
    raw = "--raw" in argv
    argv = [a for a in argv if a != "--raw"]
    if len(argv) < 2:
        print(__doc__); return 2
    src = Path(argv[1]); dst = Path(argv[2]) if len(argv) > 2 else src.with_suffix(".backup.json")
    data = convert_raw(src) if raw else convert(src)
    dst.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    if raw:
        n_sales = sum(len(v) for v in data["sales"].values()); n_ads = sum(len(v) for v in data["ads"].values())
        print(f"[원자료] 옵션 {len(data['options'])}개, 판매 {n_sales}행, 광고 {n_ads}행 → {dst}")
    else:
        n = sum(len(v) for v in data["legacy"].values()); dates = sorted(data["legacy"])
        print(f"[4번 시트] 캠페인×날짜 {n}칸, 기간 {dates[0]}~{dates[-1]}, 옵션 {len(data['options'])}개, 마진 {len(data['margins'])}개 → {dst} ({dst.stat().st_size/1e6:.1f} MB)")
    return 0


if __name__ == "__main__":
    sys.exit(_main_wrapper(sys.argv))
