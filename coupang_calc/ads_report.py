"""광고센터 > 광고 관리 > 매출 성장 > '어제' 캠페인 표 데이터 정규화.

입력 경로 3가지 모두 같은 AdsRow 로 바뀐다.
  1) collector 가 화면 표에서 읽은 dict 목록
  2) 사용자가 직접 채운 CSV/XLSX (data/inbox/ads_YYYY-MM-DD.csv)
  3) `ads-entry` 명령의 대화식 입력
"""
from __future__ import annotations

import csv
import datetime as dt
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from .common import first_match, norm_header, parse_number, parse_percent, parse_ratio

# (필드, 별칭, 파서). 별칭은 광고센터 표 헤더와 원본 시트2 헤더를 모두 포함.
ADS_FIELDS = [
    ("campaign", ["캠페인 이름", "캠페인명", "캠페인", "광고캠페인"], None),
    ("target_roas", ["목표효율", "목표 광고수익률", "목표 ROAS", "목표수익률"], parse_ratio),
    ("budget", ["광고예산", "일 예산", "일예산", "예산"], parse_number),
    ("spend", ["집행 광고비", "집행광고비", "광고비", "광고 비용", "비용"], parse_number),
    ("ad_revenue", ["광고전환 매출", "광고전환매출", "전환 매출", "전환매출", "광고 매출", "총 전환 매출"], parse_number),
    ("conversion", ["전환율", "구매전환율", "구매 전환율"], parse_percent),
    ("ctr", ["클릭률", "CTR"], parse_percent),
    ("impressions", ["노출수", "노출 수", "노출"], parse_number),
    ("clicks", ["클릭수", "클릭 수", "클릭"], parse_number),
    ("ad_orders", ["광고전환 판매수", "광고전환판매수", "총 판매수량", "판매수량", "전환수", "판매 수"], parse_number),
    ("action", ["ACTION", "액션", "메모", "비고"], None),
]
NUMERIC_FIELDS = [f for f, _, p in ADS_FIELDS if p is not None]


@dataclass
class AdsRow:
    date: dt.date
    campaign: str
    target_roas: float = 0.0  # 배수 (4.0 = 400%)
    budget: float = 0.0
    spend: float = 0.0
    ad_revenue: float = 0.0
    conversion: float = 0.0  # 분수
    ctr: float = 0.0  # 분수
    impressions: float = 0.0
    clicks: float = 0.0
    ad_orders: float = 0.0
    action: str = ""

    # 시트2 의 자동 계산 열과 같은 값 (DB 에도 저장해 두면 조회가 편하다)
    @property
    def spend_with_vat(self) -> float:
        return self.spend * 1.1

    @property
    def roas(self) -> float:
        return self.ad_revenue / self.spend if self.spend else 0.0

    @property
    def budget_usage(self) -> float:
        return self.spend / self.budget if self.budget else 0.0

    @property
    def cpc(self) -> float:
        return self.spend / self.clicks if self.clicks else 0.0

    def as_dict(self):
        d = asdict(self)
        d["date"] = self.date.isoformat()
        return d


def normalize_records(records: Iterable[Dict[str, Any]], date: dt.date) -> List[AdsRow]:
    """헤더 이름이 제각각인 dict 목록을 AdsRow 로. 캠페인명이 비어 있으면 건너뜀."""
    out: List[AdsRow] = []
    for rec in records:
        keys = list(rec.keys())
        normed = [norm_header(k) for k in keys]
        values: Dict[str, Any] = {}
        for field, aliases, parser in ADS_FIELDS:
            idx = first_match(normed, aliases)
            raw = rec[keys[idx]] if idx is not None else None
            if parser is None:
                values[field] = "" if raw is None else str(raw).strip()
            else:
                values[field] = parser(raw) or 0.0
        if not values["campaign"] or norm_header(values["campaign"]) in ("합계", "총계", "전체"):
            continue
        out.append(AdsRow(date=date, **values))
    return out


def parse_ads_file(path: Path, date: dt.date) -> List[AdsRow]:
    path = Path(path)
    if path.suffix.lower() in (".csv", ".txt"):
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
    else:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        rows = [list(r) for r in wb.worksheets[0].iter_rows(values_only=True)]
        wb.close()
    if not rows:
        return []
    h = _find_header_row(rows)
    headers = [str(c) if c is not None else "" for c in rows[h]]
    records = []
    for raw in rows[h + 1 :]:
        if raw is None or not any(c not in (None, "") for c in raw):
            continue
        records.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
    return normalize_records(records, date)


def _find_header_row(rows) -> int:
    for i, row in enumerate(rows[:20]):
        joined = "|".join(norm_header(c) for c in row)
        if "캠페인" in joined and ("광고비" in joined or "노출" in joined or "예산" in joined):
            return i
    return 0


def write_ads_template(path: Path, campaigns: Iterable[str], date: dt.date) -> Path:
    """직접 기입용 CSV 템플릿. 광고센터 표의 숫자를 그대로 옮겨 적으면 된다."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "캠페인 이름", "목표효율", "광고예산", "집행 광고비", "광고전환 매출",
                    "전환율", "클릭률", "노출수", "클릭수", "광고전환 판매수", "ACTION"])
        for c in campaigns:
            w.writerow([date.isoformat(), c, "", "", "", "", "", "", "", "", "", ""])
    return path


def prompt_ads_rows(campaigns: Iterable[str], date: dt.date, input_fn=input) -> List[AdsRow]:
    """터미널 대화식 입력. 빈 값은 0. 'skip' 을 치면 해당 캠페인은 건너뜀."""
    prompts = [
        ("target_roas", "목표효율 (예: 400% 또는 4)"),
        ("budget", "광고예산"),
        ("spend", "집행 광고비"),
        ("ad_revenue", "광고전환 매출"),
        ("conversion", "전환율 (예: 7.8%)"),
        ("ctr", "클릭률 (예: 0.15%)"),
        ("impressions", "노출수"),
        ("clicks", "클릭수"),
        ("ad_orders", "광고전환 판매수"),
        ("action", "ACTION 메모 (없으면 Enter)"),
    ]
    parsers = {f: p for f, _, p in ADS_FIELDS}
    out: List[AdsRow] = []
    for camp in campaigns:
        print(f"\n[{date}] 캠페인: {camp}  (skip 입력 시 건너뜀)")
        values: Dict[str, Any] = {}
        skipped = False
        for field, label in prompts:
            raw = input_fn(f"  {label}: ").strip()
            if raw.lower() == "skip":
                skipped = True
                break
            parser = parsers[field]
            values[field] = raw if parser is None else (parser(raw) or 0.0)
        if not skipped:
            out.append(AdsRow(date=date, campaign=camp, **values))
    return out
