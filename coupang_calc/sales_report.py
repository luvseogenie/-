"""판매자센터 > 비즈니스 인사이트 > 판매분석 > '상품별 판매 리포트' 엑셀 파서.

원본 장부 시트3 에 붙여넣던 A~M 열 13개 항목을 그대로 추출한다.
헤더 이름으로 열을 찾고, 못 찾으면 앞에서부터 13개 열을 순서대로 사용한다.
"""
from __future__ import annotations

import csv
import datetime as dt
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, List, Optional, Sequence

from .common import first_match, norm_header, parse_number, parse_percent

# (필드명, 별칭 목록) — 순서는 리포트 A~M 열 순서와 같다.
SALES_FIELDS = [
    ("option_id", ["옵션ID", "옵션 ID", "옵션아이디", "vendorItemId"]),
    ("option_name", ["옵션명", "옵션 이름"]),
    ("product_name", ["상품명", "상품 이름"]),
    ("product_id", ["등록상품ID", "등록상품 ID", "상품ID", "productId"]),
    ("category", ["카테고리"]),
    ("sales_type", ["판매방식", "판매 방식", "판매유형"]),
    ("revenue", ["매출", "매출액", "결제금액"]),
    ("orders", ["주문", "주문수", "주문 수"]),
    ("quantity", ["판매량", "판매수량", "판매 수량"]),
    ("visitors", ["방문자", "방문자수", "방문자 수"]),
    ("views", ["조회", "조회수", "상품조회"]),
    ("carts", ["장바구니", "장바구니수"]),
    ("conversion", ["구매전환율", "구매 전환율", "전환율"]),
]


@dataclass
class SalesRow:
    date: dt.date
    option_id: str
    option_name: str
    product_name: str
    product_id: str
    category: str
    sales_type: str
    revenue: float
    orders: float
    quantity: float
    visitors: float
    views: float
    carts: float
    conversion: Optional[float]  # 분수 (0.0476 = 4.76%)

    def as_dict(self):
        d = asdict(self)
        d["date"] = self.date.isoformat()
        return d


def _read_table(path: Path) -> List[List[Any]]:
    suffix = path.suffix.lower()
    if suffix in (".csv", ".txt"):
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _find_header_row(rows: Sequence[Sequence[Any]]) -> int:
    """'옵션' 과 '매출' 이 함께 있는 첫 행을 헤더로 본다. 헤더가 없으면 -1 (전부 데이터)."""
    for i, row in enumerate(rows[:20]):
        normed = [norm_header(c) for c in row]
        joined = "|".join(normed)
        if "옵션" in joined and "매출" in joined:
            return i
    return -1


def parse_sales_report(path: Path, date: dt.date) -> List[SalesRow]:
    rows = _read_table(Path(path))
    if not rows:
        return []
    h = _find_header_row(rows)
    headers = [norm_header(c) for c in rows[h]] if h >= 0 else []
    col_idx = {}
    for i, (field, aliases) in enumerate(SALES_FIELDS):
        idx = first_match(headers, aliases)
        col_idx[field] = idx if idx is not None else i  # 헤더 없으면 위치 기반

    out: List[SalesRow] = []
    for raw in rows[h + 1 :]:
        if raw is None or not any(c not in (None, "") for c in raw):
            continue

        def get(field):
            i = col_idx[field]
            return raw[i] if i < len(raw) else None

        option_id = _clean_id(get("option_id"))
        if not option_id:
            continue
        if norm_header(get("option_name")) in ("합계", "총계", "total"):
            continue
        out.append(
            SalesRow(
                date=date,
                option_id=option_id,
                option_name=_s(get("option_name")),
                product_name=_s(get("product_name")),
                product_id=_clean_id(get("product_id")),
                category=_s(get("category")),
                sales_type=_s(get("sales_type")),
                revenue=parse_number(get("revenue")) or 0.0,
                orders=parse_number(get("orders")) or 0.0,
                quantity=parse_number(get("quantity")) or 0.0,
                visitors=parse_number(get("visitors")) or 0.0,
                views=parse_number(get("views")) or 0.0,
                carts=parse_number(get("carts")) or 0.0,
                conversion=parse_percent(get("conversion")),
            )
        )
    return out


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _clean_id(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s
